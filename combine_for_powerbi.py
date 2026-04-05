import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Read all 4 CSVs
orders    = pd.read_csv("power_bi_exports/orders_full.csv")
daily     = pd.read_csv("power_bi_exports/daily_sales.csv")
revenue   = pd.read_csv("power_bi_exports/revenue_by_category.csv")
customers = pd.read_csv("power_bi_exports/customer_summary.csv")

# Write to Excel
with pd.ExcelWriter("power_bi_exports/ecommerce_dashboard.xlsx", engine="openpyxl") as writer:
    orders.to_excel(writer,    sheet_name="Orders",           index=False)
    daily.to_excel(writer,     sheet_name="Daily_Sales",      index=False)
    revenue.to_excel(writer,   sheet_name="Revenue_Category", index=False)
    customers.to_excel(writer, sheet_name="Customers",        index=False)

# Add proper table formatting to each sheet
wb = load_workbook("power_bi_exports/ecommerce_dashboard.xlsx")

def make_table(ws, name):
    max_col = get_column_letter(ws.max_column)
    max_row = ws.max_row
    ref = f"A1:{max_col}{max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)

make_table(wb["Orders"],           "Orders")
make_table(wb["Daily_Sales"],      "Daily_Sales")
make_table(wb["Revenue_Category"], "Revenue_Category")
make_table(wb["Customers"],        "Customers")

wb.save("power_bi_exports/ecommerce_dashboard.xlsx")

print("✓ Excel file created with proper tables!")
print(f"  Orders:    {len(orders)} rows")
print(f"  Daily:     {len(daily)} rows")
print(f"  Revenue:   {len(revenue)} rows")
print(f"  Customers: {len(customers)} rows")
